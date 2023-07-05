from openpyxl import load_workbook
workbook = load_workbook("test_1.xlsx")



n = 0
active_list = []

def add_input():
    # input the table number
    table_num = int(input("Input the table number: "))
    active_list.append(table_num)

    # input the order as string then add it to the order list
    order = input("add the order: ")
    active_list.append(order)

    # check if there are other orders for the same table
    add_order = input("do u want to add another order? y/n: ")

    # while loop for adding orders to the same table
    while add_order == "y":
        order = input("add the order: ")

        # creating the order_list
        active_list.append(order)
        add_order = input("do u want to add another order? y/n: ")




def active_tabs_input():
    add_input()
    worksheet = workbook['active_tabs']
    num_rows = worksheet.max_row
    for i, item in enumerate(active_list, start=0):
        worksheet.cell(row=num_rows + 1 , column=i + 1).value = item
    workbook.save("test_1.xlsx")
    active_list.clear()

while True:
    active_tabs_input()